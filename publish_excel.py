# -*- coding: utf-8 -*-
"""
상담 공지 엑셀 → 구글 드라이브 업로드 → 슬랙에 다운로드 링크 공지

흐름:
  1. consult_excel.generate() 로 지점별 엑셀 생성
  2. 내 드라이브 '상담공지_엑셀/YYYY-MM-DD' 폴더에 업로드
  3. 링크 공유(보기) 설정 후 슬랙 webhook으로 링크 목록 전송

인증: ~/.clasprc.json 의 구글 OAuth 토큰 (drive.file 권한 — 이 앱이 만든 파일만 접근)

실행:
  py -X utf8 publish_excel.py            # 생성+업로드+슬랙 공지
  py -X utf8 publish_excel.py --dry-run  # 슬랙 전송 없이 링크만 출력
"""
from __future__ import annotations

import argparse
import io
import json
import mimetypes
import os
import sys
import time
import urllib.parse
import urllib.request
import uuid
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

try:
    import keyring
except ImportError:
    keyring = None

import consult_excel
import consult_report as cr

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
FOLDER_MIME = "application/vnd.google-apps.folder"
RC = os.path.join(os.path.expanduser("~"), ".clasprc.json")
ROOT_FOLDER = "상담공지_엑셀"


# ---------- 구글 OAuth (clasp 토큰 재사용) ----------
def google_token() -> str:
    env = os.environ.get("GOOGLE_OAUTH_JSON")  # GitHub Actions용: .clasprc.json 내용
    store = json.loads(env) if env else json.loads(open(RC, encoding="utf-8").read())
    cred = store["tokens"]["default"]
    if not env and cred.get("expiry_date", 0) > time.time() * 1000 + 60000:
        return cred["access_token"]
    data = urllib.parse.urlencode({
        "client_id": cred["client_id"],
        "client_secret": cred["client_secret"],
        "refresh_token": cred["refresh_token"],
        "grant_type": "refresh_token",
    }).encode()
    req = urllib.request.Request("https://oauth2.googleapis.com/token", data=data)
    with urllib.request.urlopen(req, timeout=30) as res:
        tok = json.loads(res.read().decode())
    if not env:
        cred["access_token"] = tok["access_token"]
        cred["expiry_date"] = int(time.time() * 1000) + tok.get("expires_in", 3600) * 1000
        open(RC, "w", encoding="utf-8").write(json.dumps(store, indent=2))
    return tok["access_token"]


# ---------- Drive API ----------
def drive(token: str, method: str, path: str, body: dict | None = None, query: dict | None = None) -> dict:
    url = f"https://www.googleapis.com/drive/v3{path}"
    if query:
        url += "?" + urllib.parse.urlencode(query)
    data = json.dumps(body).encode() if body is not None else None
    headers = {"Authorization": f"Bearer {token}"}
    if data:
        headers["Content-Type"] = "application/json"
    req = urllib.request.Request(url, data=data, method=method, headers=headers)
    with urllib.request.urlopen(req, timeout=60) as res:
        raw = res.read().decode()
        return json.loads(raw) if raw else {}


def find_or_create_folder(token: str, name: str, parent: str | None = None) -> str:
    q = f"name = '{name}' and mimeType = '{FOLDER_MIME}' and trashed = false"
    if parent:
        q += f" and '{parent}' in parents"
    found = drive(token, "GET", "/files", query={"q": q, "fields": "files(id)"})
    if found.get("files"):
        return found["files"][0]["id"]
    meta = {"name": name, "mimeType": FOLDER_MIME}
    if parent:
        meta["parents"] = [parent]
    return drive(token, "POST", "/files", body=meta)["id"]


def upload_file(token: str, path: Path, folder_id: str, drive_name: str) -> dict:
    """같은 이름 파일이 있으면 내용 덮어쓰기(링크 유지), 없으면 새로 만들고 링크공유(뷰어) 설정."""
    q = (f"name = '{drive_name}' and '{folder_id}' in parents and trashed = false")
    found = drive(token, "GET", "/files", query={"q": q, "fields": "files(id,webViewLink)"})
    content = path.read_bytes()

    if found.get("files"):
        # 기존 파일 초기화: 내용만 교체 → 파일 ID·링크·권한 그대로 유지
        fid = found["files"][0]["id"]
        req = urllib.request.Request(
            f"https://www.googleapis.com/upload/drive/v3/files/{fid}?uploadType=media&fields=id,webViewLink",
            data=content, method="PATCH",
            headers={"Authorization": f"Bearer {token}", "Content-Type": XLSX_MIME},
        )
        with urllib.request.urlopen(req, timeout=120) as res:
            info = json.loads(res.read().decode())
        _ensure_anyone_writer(token, fid)  # 지점이 '제외(O)' 체크하도록 편집권한
        return info

    boundary = uuid.uuid4().hex
    meta = json.dumps({"name": drive_name, "parents": [folder_id]}).encode()
    body = b"".join([
        f"--{boundary}\r\nContent-Type: application/json; charset=UTF-8\r\n\r\n".encode(),
        meta,
        f"\r\n--{boundary}\r\nContent-Type: {XLSX_MIME}\r\n\r\n".encode(),
        content,
        f"\r\n--{boundary}--".encode(),
    ])
    req = urllib.request.Request(
        "https://www.googleapis.com/upload/drive/v3/files?uploadType=multipart&fields=id,webViewLink",
        data=body, method="POST",
        headers={"Authorization": f"Bearer {token}",
                 "Content-Type": f"multipart/related; boundary={boundary}"},
    )
    with urllib.request.urlopen(req, timeout=120) as res:
        info = json.loads(res.read().decode())
    _ensure_anyone_writer(token, info["id"])  # 링크 있는 사용자 편집 가능(지점 '제외(O)' 표시용)
    return info


# ---------- 슬랙 ----------
def send_slack(payload: dict | str) -> None:
    if isinstance(payload, str):
        payload = {"text": payload}
    # 1순위: 봇(chat.postMessage) — 발신자 이름을 '지점관리'로 지정 (chat:write.customize 필요)
    bot = (os.environ.get("SLACK_BOT_TOKEN")
           or (keyring.get_password("carefor-auto", "slack_bot_token") if keyring else None))
    channel = os.environ.get("CONSULT_CHANNEL") or "C0BC37EB38C"  # 프로그램관리 채널
    if bot and channel:
        p = dict(payload)
        p["channel"] = channel
        p.setdefault("username", "지점관리")
        p.setdefault("icon_emoji", ":office:")
        body = json.dumps(p).encode("utf-8")
        req = urllib.request.Request(
            "https://slack.com/api/chat.postMessage", data=body,
            headers={"Content-Type": "application/json; charset=utf-8",
                     "Authorization": f"Bearer {bot}"})
        out = json.loads(urllib.request.urlopen(req, timeout=30).read().decode("utf-8"))
        if not out.get("ok"):
            raise SystemExit(f"슬랙(봇) 전송 실패: {out.get('error')} — {out}")
        return
    # 폴백: 기존 웹훅 (봇 토큰 없을 때)
    hook = (os.environ.get("ARONGI_WEBHOOK_URL")
            or (keyring.get_password("carefor-auto", "arongi_webhook_url") if keyring else None)
            or os.environ.get("SLACK_WEBHOOK_URL")
            or (keyring.get_password("carefor-auto", "slack_webhook_url") if keyring else None))
    if not hook:
        raise SystemExit("발송 자격증명(봇 토큰/웹훅) 없음.")
    body = json.dumps(payload).encode("utf-8")
    req = urllib.request.Request(hook, data=body,
                                 headers={"Content-Type": "application/json; charset=utf-8"})
    out = urllib.request.urlopen(req, timeout=30).read().decode("utf-8")
    if out.strip() != "ok":
        raise SystemExit(f"슬랙 전송 실패: {out}")


def resolve_mention_ids(names: list) -> dict:
    """이름 목록 → 슬랙 user ID 매핑 (봇 토큰으로 users.list 조회). 못 찾은 이름은 결과에서 빠짐."""
    tok = (os.environ.get("SLACK_BOT_TOKEN")
           or (keyring.get_password("carefor-auto", "slack_bot_token") if keyring else None))
    if not tok or not names:
        return {}
    import urllib.parse
    idmap = {}
    cursor = ""
    while True:
        url = "https://slack.com/api/users.list?limit=200"
        if cursor:
            url += "&cursor=" + urllib.parse.quote(cursor)
        req = urllib.request.Request(url, headers={"Authorization": f"Bearer {tok}"})
        d = json.loads(urllib.request.urlopen(req, timeout=30).read().decode("utf-8"))
        if not d.get("ok"):
            print(f"[경고] users.list 실패: {d.get('error')}")
            break
        for m in d.get("members", []):
            if m.get("deleted") or m.get("is_bot"):
                continue
            prof = m.get("profile", {})
            cand = [(m.get("real_name") or ""), (prof.get("real_name") or ""),
                    (prof.get("display_name") or ""), (prof.get("real_name_normalized") or ""),
                    (prof.get("display_name_normalized") or "")]
            cand = [c.strip() for c in cand]
            for nm in names:
                if nm in idmap:
                    continue
                if any(c == nm for c in cand) or any(nm in c for c in cand if c):
                    idmap[nm] = m["id"]
        cursor = d.get("response_metadata", {}).get("next_cursor", "")
        if not cursor:
            break
    return idmap


def _download(token: str, file_id: str) -> bytes:
    req = urllib.request.Request(
        f"https://www.googleapis.com/drive/v3/files/{file_id}?alt=media",
        headers={"Authorization": f"Bearer {token}"})
    with urllib.request.urlopen(req, timeout=120) as res:
        return res.read()


def _ensure_anyone_writer(token: str, fid: str) -> None:
    """지점이 '제외(O)' 체크할 수 있게 링크 편집권한(anyone writer) 보장."""
    perms = drive(token, "GET", f"/files/{fid}/permissions", query={"fields": "permissions(id,type,role)"})
    anyone = next((p for p in perms.get("permissions", []) if p.get("type") == "anyone"), None)
    if anyone and anyone.get("role") != "writer":
        drive(token, "PATCH", f"/files/{fid}/permissions/{anyone['id']}", body={"role": "writer"})
    elif not anyone:
        drive(token, "POST", f"/files/{fid}/permissions", body={"type": "anyone", "role": "writer"})


def harvest_exclude_marks(token: str, folder_id: str) -> list:
    """지점 상담공지 엑셀 '신규상담 미입력' 시트의 '제외(O)' 표시된 행 번호 수집 → [(번호,사유)]."""
    from openpyxl import load_workbook
    names = [f"{full}_상담공지.xlsx" for _, full in cr.CENTER_ORDER] + ["전체_상담공지.xlsx"]
    seen, pairs = set(), []
    for name in names:
        found = drive(token, "GET", "/files",
                      query={"q": f"name = '{name}' and '{folder_id}' in parents and trashed = false",
                             "fields": "files(id)"})
        if not found.get("files"):
            continue
        try:
            wb = load_workbook(io.BytesIO(_download(token, found["files"][0]["id"])),
                               read_only=True, data_only=True)
        except Exception as e:
            print(f"[제외수집] {name} 열기 실패: {e}")
            continue
        try:
            if "신규상담 미입력" not in wb.sheetnames:
                continue
            rows = list(wb["신규상담 미입력"].iter_rows(values_only=True))
        finally:
            wb.close()
        if not rows:
            continue
        header = [str(h or "") for h in rows[0]]
        ci_ex = next((i for i, h in enumerate(header) if h.startswith("제외")), None)
        ci_ph = next((i for i, h in enumerate(header) if "번호" in h), None)
        if ci_ex is None or ci_ph is None:
            continue
        for row in rows[1:]:
            if len(row) <= max(ci_ex, ci_ph):
                continue
            if str(row[ci_ex] or "").strip() and str(row[ci_ph] or "").strip():
                d = cr._norm_phone(row[ci_ph])
                if len(d) >= 10 and d not in seen:
                    seen.add(d)
                    pairs.append((str(row[ci_ph]).strip(), f"지점 제외표시({name.split('_')[0]})"))
    return pairs


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    today = date.today()
    token = google_token()
    root_id = find_or_create_folder(token, ROOT_FOLDER)

    # 지점 엑셀의 '제외(O)' 표시 수집 → 제외번호 자동 등록 (재생성 전에!)
    try:
        marks = harvest_exclude_marks(token, root_id)
        if marks:
            import excl_store
            n = excl_store.add_phones(marks)
            print(f"지점 '제외(O)' 표시 {len(marks)}건 확인 → 제외번호 {n}건 신규 추가")
    except Exception as e:
        print(f"[경고] 지점 제외표시 수집 실패(무시하고 진행): {e}")

    out_dir, paths, summaries = consult_excel.generate(today)

    links = []
    for p in paths:
        # 드라이브에는 날짜 없는 고정 이름으로 → 매번 같은 파일에 덮어쓰기 (링크 불변)
        center = p.stem.split("_")[0]
        drive_name = f"{center}_상담공지.xlsx"
        info = upload_file(token, p, root_id, drive_name)
        links.append((center, info["webViewLink"]))
        print(f"업로드(덮어쓰기): {drive_name}")

    weekday = "월화수목금토일"[today.weekday()]
    link_map = dict(links)

    # 집계표 (센터별 요약 + 상담 대기 줄 포함) — 한 페이지 통합 공지
    import consult_report as cr
    shorts = [short for short, _ in cr.CENTER_ORDER]
    smap = {s["center"]: s for s in summaries}
    cols = [smap[full] for _, full in cr.CENTER_ORDER]
    LABEL_W = 16
    col_ws = [max(cr._w(n), 4) + 2 for n in shorts]
    table_lines = [
        cr._rpad("", LABEL_W) + "".join(cr._lpad(n, w) for n, w in zip(shorts, col_ws)),
        "─" * (LABEL_W + sum(col_ws)),
    ]
    for label, key in [("신규상담(누적)", "total"), ("시트 미입력", "miss"),
                       ("미입력률", "rate"), ("대기(아웃콜)", "wait")]:
        table_lines.append(cr._rpad(label, LABEL_W) + "".join(
            cr._lpad(str(s[key]), w) for s, w in zip(cols, col_ws)))
    table = "\n".join(table_lines)

    # 지점별 엑셀 링크 한 줄
    link_parts = [f"<{link_map.get(full, '')}|{short}>" for short, full in cr.CENTER_ORDER]
    link_line = "📎 *상세 명단(엑셀)*:  " + "  ·  ".join(link_parts) \
                + f"  ·  <{link_map.get('전체', '')}|전체>"

    msg = {
        "text": f"☎️ 신규상담 시트 입력 현황(아롱이) {today.strftime('%Y.%m.%d')}({weekday})",
        "blocks": [
            {"type": "header", "text": {"type": "plain_text",
                "text": "☎️ 신규상담 시트 입력 현황(아롱이)", "emoji": True}},
            {"type": "context", "elements": [{"type": "mrkdwn",
                "text": f"{today.strftime('%Y.%m.%d')}({weekday}) · 전일자 기준 · 2026년 5월~ 누적"}]},
            {"type": "section", "text": {"type": "mrkdwn",
                "text": "💛 상담 한 분 한 분이 소중한 인연입니다 — 오늘의 상담 한 통이 어르신과의 첫 만남이 됩니다."}},
            {"type": "section", "text": {"type": "mrkdwn", "text": f"```\n{table}\n```"}},
            {"type": "section", "text": {"type": "mrkdwn", "text": link_line}},
            {"type": "context", "elements": [{"type": "mrkdwn",
                "text": "📝 상담시트 입력 부탁드립니다. 연락처·아웃콜 차수 등 상세는 엑셀(링크 고정) 참조.\n"
                        "주간보호가 아닌 건은 엑셀 맨 뒤 '제외 ✔' 칸을 체크하시면 다음날 자동 제외됩니다.\n"
                        "상담 시트 관련 문의 사항이 있으시면 본부 매니저에게 문의 부탁드립니다. 감사합니다. 🙏\n"
                        "🗓️ 이 공지는 매주 화·목요일 오전 9시 30분에 발송됩니다."}]},
        ],
    }

    # 센터장 태그 (CONSULT_MENTIONS: 콤마구분 이름) — 봇 토큰으로 ID 조회 후 상단 멘션 블록 삽입
    mnames = [x.strip() for x in os.environ.get("CONSULT_MENTIONS", "").split(",") if x.strip()]
    if mnames:
        idmap = resolve_mention_ids(mnames)
        found = [idmap[n] for n in mnames if n in idmap]
        missing = [n for n in mnames if n not in idmap]
        print(f"태그 조회 결과: " + ", ".join(f"{n}={idmap.get(n,'못찾음')}" for n in mnames))
        if missing:
            print(f"[경고] 태그 ID 못 찾은 이름: {missing}")
        if found:
            mtext = " ".join(f"<@{i}>" for i in found)
            msg["text"] = mtext + " " + msg["text"]
            msg["blocks"].insert(0, {"type": "section", "text": {"type": "mrkdwn", "text": mtext}})

    print("--- 슬랙 메시지 ---")
    print(json.dumps(msg, ensure_ascii=False)[:1500])
    if args.dry_run:
        print("(dry-run: 전송 안 함)")
        return
    send_slack(msg)
    print("전송 완료 → #차량관리 (webhook)")


if __name__ == "__main__":
    main()
