# -*- coding: utf-8 -*-
"""지점점검표 엑셀·PDF 내보내기 (더블클릭: 점검표_내보내기.bat).

데이터:
  · 자동판정 — audit_results/*.json (정기/수동 점검 결과)
  · 채점 점수·메모 — 점수 공유 웹훅 (대시보드에서 📤 본부 공유 클릭한 것)
    → 공유 안 한 지점은 자동판정 양호분만 자동 채점으로 합성

출력: 점검표/지점점검표_YYYYMMDD.xlsx + .pdf
"""
from __future__ import annotations

import json
import sys
import urllib.request
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
BASE = Path(__file__).resolve().parent
sys.path.insert(0, str(BASE))

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from audit.items import ITEMS

SCORES_HOOK = ("https://script.google.com/macros/s/AKfycbxCaMyM26xaLNlYUof-Jxac88jfWggzphLDvBEIRlDY-2Bn8S9wF5HWt52QupXWkxlO/"
               "exec?token=audit-scores-2026-cheongju")
BRANCHES = ["둔산점", "서구점", "천안점", "청주 오창점"]
AUDIT_DIR = BASE / "audit_results"
OUT_DIR = BASE / "점검표"

H_FILL = PatternFill("solid", fgColor="2F5496")
H_FONT = Font(bold=True, color="FFFFFF")
OK_FILL = PatternFill("solid", fgColor="E2F0E5")
BAD_FILL = PatternFill("solid", fgColor="FDECEC")
NA_FILL = PatternFill("solid", fgColor="EFEFEF")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def load_data():
    autos = {}
    for f in AUDIT_DIR.glob("*.json"):
        try:
            d = json.loads(f.read_text(encoding="utf-8"))
            autos[d["branch"]] = d
        except Exception:
            continue
    try:
        scores = json.loads(urllib.request.urlopen(SCORES_HOOK, timeout=60).read().decode()).get("scores", {})
    except Exception as e:
        print(f"⚠️ 점수 웹훅 조회 실패({e}) — 자동판정만으로 생성")
        scores = {}
    return autos, scores


def eff_sub_value(it, sub, uploaded, auto_d):
    """기준별 유효 점수: 업로드값 > 자동판정 양호 합성 > 미입력."""
    no = str(it["no"])
    up = (uploaded or {}).get("sub_values", {}).get(no, {}).get(sub["label"])
    if up is not None:
        return up["v"], ("자동" if up.get("auto") else "수기")
    r = (auto_d or {}).get("item_results", {}).get(no)
    if r and it.get("auto_subs") and sub["label"] in it["auto_subs"]:
        ss = r.get("sub_status") or {}
        st = ss.get(sub["label"], r.get("status"))
        if st == "양호":
            return str(sub["score"]), "자동"
    return None, ""


def status_of(v, sub):
    if v is None:
        return "미입력"
    if v == "NA":
        return "해당없음"
    x = float(v)
    if x >= sub["score"]:
        return "충족"
    return "미충족" if x == 0 else "부분충족"


def build_excel(autos, scores, today, branches=None) -> Path:
    branches = branches or BRANCHES
    wb = Workbook()
    ws = wb.active
    ws.title = "요약"
    ws.append(["지점", "총점(100점)", "기준 입력", "해당없음 배점", "자동판정 양호", "자동판정 미흡", "자동수집 시각", "기관 지정일"])
    for c in ws[1]:
        c.fill, c.font, c.alignment = H_FILL, H_FONT, CENTER

    branch_rows = {}
    for b in branches:
        d = autos.get(b)
        up = scores.get(b)
        earned = denom = filled = total = na_pts = 0
        rows = []
        n_ok = n_bad = 0
        if d:
            for r in d.get("item_results", {}).values():
                if r.get("status") == "양호":
                    n_ok += 1
                else:
                    n_bad += 1
        for it in ITEMS:
            for sub in it.get("subs", []):
                v, src = eff_sub_value(it, sub, up, d)
                st = status_of(v, sub)
                total += 1
                if v is not None:
                    filled += 1
                    if v == "NA":
                        na_pts += sub["score"]
                    else:
                        earned += float(v)
                        denom += sub["score"]
                auto_r = (d or {}).get("item_results", {}).get(str(it["no"]), {})
                memo = (up or {}).get("memos", {}).get(str(it["no"]), "")
                rows.append([it["no"], it["name"], sub["label"], sub["text"], sub["score"],
                             ("" if v is None else ("해당없음" if v == "NA" else float(v))), st, src,
                             auto_r.get("status", ""), memo])
        branch_rows[b] = rows
        ws.append([b, round(earned, 2), f"{filled}/{total}", round(na_pts, 2), n_ok, n_bad,
                   (d or {}).get("run_at", "-"), (d or {}).get("opened") or "-"])

    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = CENTER
    for i, w in enumerate([14, 13, 11, 13, 13, 13, 17, 13], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    headers = ["번호", "항목", "기준", "기준 내용", "배점", "점수", "판정", "채점", "자동판정", "메모"]
    widths = [7, 18, 7, 55, 7, 9, 10, 7, 10, 28]
    for b in branches:
        wsb = wb.create_sheet(b)
        wsb.append(headers)
        for c in wsb[1]:
            c.fill, c.font, c.alignment = H_FILL, H_FONT, CENTER
        prev_no = None
        for row in branch_rows[b]:
            disp = list(row)
            if disp[0] == prev_no:
                disp[0], disp[1], disp[9] = "", "", ""
            else:
                prev_no = disp[0]
            wsb.append(disp)
            st = row[6]
            fill = OK_FILL if st in ("충족",) else (BAD_FILL if st in ("미충족", "부분충족") else (NA_FILL if st == "해당없음" else None))
            for j, c in enumerate(wsb[wsb.max_row], 1):
                c.alignment = LEFT if j in (4, 10) else CENTER
                if fill and j in (6, 7):
                    c.fill = fill
        for i, w in enumerate(widths, 1):
            wsb.column_dimensions[get_column_letter(i)].width = w
        wsb.freeze_panes = "A2"
        wsb.auto_filter.ref = wsb.dimensions

    tag = branches[0] + "_" if len(branches) == 1 else ""
    p = OUT_DIR / f"지점점검표_{tag}{today:%Y%m%d}.xlsx"
    wb.save(p)
    return p


def build_pdf(autos, scores, today, branches=None) -> Path:
    branches = branches or BRANCHES
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    from reportlab.platypus import (Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle, PageBreak)
    from reportlab.lib.styles import ParagraphStyle

    pdfmetrics.registerFont(TTFont("Malgun", r"C:\Windows\Fonts\malgun.ttf"))
    pdfmetrics.registerFont(TTFont("MalgunB", r"C:\Windows\Fonts\malgunbd.ttf"))
    h1 = ParagraphStyle("h1", fontName="MalgunB", fontSize=15, spaceAfter=4)
    meta = ParagraphStyle("meta", fontName="Malgun", fontSize=8.5, textColor=colors.grey, spaceAfter=8)

    tag = branches[0] + "_" if len(branches) == 1 else ""
    p = OUT_DIR / f"지점점검표_{tag}{today:%Y%m%d}.pdf"
    doc = SimpleDocTemplate(str(p), pagesize=A4, topMargin=14 * mm, bottomMargin=12 * mm,
                            leftMargin=12 * mm, rightMargin=12 * mm)
    story = []
    for bi, b in enumerate(branches):
        d = autos.get(b)
        up = scores.get(b)
        earned = denom_all = 0.0
        na_pts = 0.0
        filled = total = 0
        data = [["번호", "항목", "배점", "득점", "판정", "자동판정"]]
        row_styles = []
        for it in ITEMS:
            it_earned = it_na = 0.0
            it_filled = 0
            for sub in it.get("subs", []):
                v, _src = eff_sub_value(it, sub, up, d)
                total += 1
                if v is not None:
                    filled += 1
                    it_filled += 1
                    if v == "NA":
                        it_na += sub["score"]
                        na_pts += sub["score"]
                    else:
                        it_earned += float(v)
                        earned += float(v)
            denom_all += it.get("total", 0)
            auto_r = (d or {}).get("item_results", {}).get(str(it["no"]), {})
            possible = it.get("total", 0) - it_na
            if it_filled == 0:
                verdict = "미입력"
            elif possible <= 0:
                verdict = "해당없음"
            elif it_earned >= possible:
                verdict = "충족"
            else:
                verdict = "미달"
            data.append([str(it["no"]), it["name"], f"{it.get('total', 0):g}",
                         ("-" if it_filled == 0 else f"{it_earned:g}"), verdict, auto_r.get("status", "")])
            i = len(data) - 1
            if verdict == "충족":
                row_styles.append(("TEXTCOLOR", (4, i), (4, i), colors.HexColor("#2c8a41")))
            elif verdict == "미달":
                row_styles.append(("TEXTCOLOR", (4, i), (4, i), colors.HexColor("#c02020")))

        story.append(Paragraph(f"지점 점검표 — {b}", h1))
        story.append(Paragraph(
            f"생성 {today:%Y-%m-%d} · 총점 <b>{earned:g}점</b> / 만점 {100 - na_pts:g}점(해당없음 {na_pts:g}점 제외)"
            f" · 기준 입력 {filled}/{total} · 자동수집 {(d or {}).get('run_at', '-')}"
            f" · 기관 지정일 {(d or {}).get('opened') or '-'}", meta))
        t = Table(data, colWidths=[11 * mm, 62 * mm, 14 * mm, 14 * mm, 18 * mm, 18 * mm], repeatRows=1)
        t.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, 0), "MalgunB"),
            ("FONTNAME", (0, 1), (-1, -1), "Malgun"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2f5496")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#c9d2e3")),
            ("ALIGN", (0, 0), (0, -1), "CENTER"),
            ("ALIGN", (2, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6fa")]),
        ] + row_styles))
        story.append(t)
        if bi < len(branches) - 1:
            story.append(PageBreak())
    doc.build(story)
    return p


def pick_branches() -> list[str]:
    print("\n내보낼 지점을 선택하세요:")
    for i, b in enumerate(BRANCHES, 1):
        print(f"  {i}. {b}")
    print(f"  {len(BRANCHES) + 1}. 전체 지점")
    sel = input("번호 입력: ").strip()
    if sel.isdigit():
        n = int(sel)
        if 1 <= n <= len(BRANCHES):
            return [BRANCHES[n - 1]]
        if n == len(BRANCHES) + 1:
            return list(BRANCHES)
    print("  잘못된 입력 — 전체 지점으로 진행합니다.")
    return list(BRANCHES)


def pick_format() -> str:
    print("\n형식을 선택하세요:")
    print("  1. 엑셀 (기준별 상세 + 색상)")
    print("  2. PDF (항목별 요약 점검표)")
    print("  3. 둘 다")
    sel = input("번호 입력: ").strip()
    return {"1": "excel", "2": "pdf"}.get(sel, "both")


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--branch", help="지점명 (미지정 시 메뉴)")
    ap.add_argument("--format", choices=["excel", "pdf", "both"], help="형식 (미지정 시 메뉴)")
    ap.add_argument("--no-open", action="store_true")
    args = ap.parse_args()

    OUT_DIR.mkdir(exist_ok=True)
    today = date.today()
    autos, scores = load_data()
    if not autos:
        print("audit_results/*.json 이 없습니다 — 먼저 점검을 실행하세요.")
        return

    if args.branch:
        branches = [b for b in BRANCHES if args.branch in b]
        if not branches:
            print(f"지점 '{args.branch}' 없음")
            return
        fmt = args.format or "both"
    else:
        branches = pick_branches()
        fmt = args.format or pick_format()

    print(f"\n▶ {', '.join(branches)} / {'엑셀+PDF' if fmt == 'both' else fmt.upper()} 생성 중...")
    if fmt in ("excel", "both"):
        xp = build_excel(autos, scores, today, branches)
        print(f"📗 엑셀: {xp.name}")
    if fmt in ("pdf", "both"):
        try:
            pp = build_pdf(autos, scores, today, branches)
            print(f"📕 PDF: {pp.name}")
        except Exception as e:
            print(f"PDF 생성 실패: {e}")
    print(f"\n완료 — 저장 위치: {OUT_DIR}")
    if not args.no_open:
        import os
        os.startfile(str(OUT_DIR))


if __name__ == "__main__":
    main()
