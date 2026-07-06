# -*- coding: utf-8 -*-
"""
상담시트 미입력 건 ↔ 케어포 수급자 대조

흐름:
  1. 케어포 4개 지점 로그인 → 1-7 수급자 현황리포트 → 연간 조회 → 엑셀 다운로드
  2. 수급자 연락처 + 보호자 휴대폰/전화번호 수집
  3. 신규상담 세부사항의 '상담시트 미입력(N)' 건 연락처와 대조
  4. 결과 엑셀 저장 (상담공지_엑셀/케어포대조_날짜.xlsx)

실행: py -X utf8 carefor_phone_check.py [--limit-branch 둔산점]
"""
from __future__ import annotations

import argparse
import re
import sys
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")
sys.path.insert(0, str(Path(__file__).resolve().parent))

import yaml
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from playwright.sync_api import sync_playwright

import consult_report as cr
from src import credentials
from src.carefor_client import build_spa_hash, extract_g_pammgno, _navigate_spa
from src.config import config_path

PORTAL_URL = "https://eform.caring.co.kr/carefor"
DN_BASE = "https://dn.carefor.co.kr/"
DL_DIR = Path(__file__).resolve().parent / "상담공지_엑셀" / "케어포_수급자현황"

# 케어포 지점명(config) ↔ 신규상담 센터명 매칭
BRANCH_TO_CENTER = {"둔산점": "대전둔산점", "서구점": "대전서구점",
                    "천안점": "천안점", "청주 오창점": "청주오창점"}


def digits(p) -> str:
    d = re.sub(r"[^0-9]", "", str(p or ""))
    if len(d) == 10 and d.startswith("10"):
        d = "0" + d
    return d


def download_branch_report(ctmnumb: str, branch: str, headless: bool = True) -> Path:
    """1-7 연간 조회 엑셀 다운로드 → 파일 경로 반환."""
    portal_id, portal_pw = credentials.get_portal_credentials()
    DL_DIR.mkdir(parents=True, exist_ok=True)
    dest = DL_DIR / f"{branch}_수급자현황_연간.xlsx"

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        ctx = browser.new_context(http_credentials={"username": portal_id, "password": portal_pw},
                                  accept_downloads=True)
        portal_page = ctx.new_page()
        print(f"[{branch}] 포털 로그인...")
        portal_page.goto(PORTAL_URL, wait_until="domcontentloaded")
        portal_page.wait_for_function("typeof login2 === 'function'", timeout=15000)
        with ctx.expect_page(timeout=60000) as npi:
            portal_page.evaluate(f"login2('{ctmnumb}')")
        page = npi.value
        page.wait_for_load_state("domcontentloaded", timeout=30000)
        page.wait_for_load_state("networkidle", timeout=30000)

        g = extract_g_pammgno(page)
        h = build_spa_hash("left_sub1", "/share/patient/view.patient_report", "1-7.수급자 현황 리포트", g)
        _navigate_spa(page, f"{DN_BASE}#{h}")
        page.wait_for_timeout(4000)

        # 조회기준 '연간' 클릭
        clicked = page.evaluate("""
          (() => {
            const el = Array.from(document.querySelectorAll('*')).find(el => {
              const own = Array.from(el.childNodes).filter(n => n.nodeType === 3)
                .map(n => n.textContent.trim()).join('');
              return own === '연간' && el.offsetParent;
            });
            if (el) { el.click(); return true; }
            return false;
          })()
        """)
        if not clicked:
            raise RuntimeError("'연간' 버튼을 찾지 못했습니다.")
        page.wait_for_timeout(4000)

        print(f"[{branch}] 엑셀 다운로드...")
        with page.expect_download(timeout=120000) as dl_info:
            page.evaluate("excel_patient_report()")
        dl_info.value.save_as(dest)
        browser.close()
    print(f"[{branch}] 저장: {dest.name}")
    return dest


def parse_report(path: Path, branch: str) -> list[dict]:
    """다운로드 엑셀 → [{name, status, start, quit, phones{연락처유형: 번호}}]"""
    wb = load_workbook(path, read_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    header_idx = next(i for i, r in enumerate(rows) if r and "수급자명" in [str(c).strip() for c in r if c])
    header = [str(c).strip() if c else "" for c in rows[header_idx]]

    def col(name):
        return header.index(name) if name in header else None

    # 보호자 휴대폰/전화번호는 뒤쪽(보호자정보 구역)에 위치 — index로 자동 탐색
    idx = {
        "name": col("수급자명"), "status": col("수급현황"),
        "start": col("급여개시일"), "quit": col("퇴소일"),
        "연락처(수급자)": col("연락처"), "휴대폰(보호자)": col("휴대폰"), "전화번호(보호자)": col("전화번호"),
    }
    out = []
    for r in rows[header_idx + 1:]:
        if not r or idx["name"] is None or not r[idx["name"]]:
            continue
        phones = {}
        for label in ["연락처(수급자)", "휴대폰(보호자)", "전화번호(보호자)"]:
            i = idx[label]
            if i is not None and r[i]:
                d = digits(r[i])
                if len(d) >= 9:
                    phones[label] = d
        out.append({
            "branch": branch,
            "name": str(r[idx["name"]]).strip(),
            "status": str(r[idx["status"]] or "").strip() if idx["status"] is not None else "",
            "start": str(r[idx["start"]] or "").strip() if idx["start"] is not None else "",
            "quit": str(r[idx["quit"]] or "").strip() if idx["quit"] is not None else "",
            "phones": phones,
        })
    return out


def build_phone_index(skip_download: bool = True, limit_branch: str | None = None) -> dict:
    """지점별 수급자 명단 확보 → 전화번호 색인. skip_download=True면 기존 파일 재사용."""
    cfg = yaml.safe_load(config_path().read_text(encoding="utf-8"))
    branches = cfg["branches"]
    if limit_branch:
        branches = [b for b in branches if b["name"] == limit_branch]

    patients = []
    for b in branches:
        path = DL_DIR / f"{b['name']}_수급자현황_연간.xlsx"
        if not (skip_download and path.exists()):
            path = download_branch_report(b["ctmnumb"], b["name"])
        patients += parse_report(path, b["name"])

    phone_index: dict[str, list[tuple[dict, str]]] = {}
    for pt in patients:
        for label, d in pt["phones"].items():
            phone_index.setdefault(d, []).append((pt, label))
    phone_index["__count__"] = len(patients)
    return phone_index


def lookup(phone_index: dict, phone: str) -> tuple[dict, str] | None:
    """미입력 건 연락처로 케어포 수급자 조회. 없으면 None."""
    hits = phone_index.get(digits(phone), [])
    return hits[0] if hits else None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--limit-branch", help="특정 지점만 (예: 둔산점)")
    ap.add_argument("--skip-download", action="store_true", help="기존 다운로드 파일 재사용")
    ap.add_argument("--no-slack", action="store_true", help="드라이브 업로드·슬랙 공지 생략")
    args = ap.parse_args()

    phone_index = build_phone_index(skip_download=args.skip_download, limit_branch=args.limit_branch)
    print(f"케어포 수급자 총 {phone_index.pop('__count__')}명 수집")

    # 2) 미입력 건 로드 + 대조
    rows = cr.load_rows_from_webhook()
    miss = [r for r in rows if r["sheet_entered"] == "N"]
    print(f"상담시트 미입력 {len(miss)}건 대조 중...")

    results = []
    for m in miss:
        d = digits(m["phone"])
        hits = phone_index.get(d, [])
        results.append((m, hits))

    matched = [x for x in results if x[1]]
    print(f"\n=== 케어포 등록 확인: {len(matched)}건 / 미입력 {len(miss)}건 ===")
    for m, hits in matched:
        pt, label = hits[0]
        print(f"  {m['center']} | 상담일 {m['consult_date']} | {m['phone']} → "
              f"{pt['branch']} {pt['name']} ({pt['status']}, {label}, 개시 {pt['start']})")

    # 3) 결과 엑셀
    today = date.today()
    wb = Workbook()
    ws = wb.active
    ws.title = "케어포 대조결과"
    ws.append(["상담 센터", "연월", "상담일자", "고객 번호", "입소 여부(시트)",
               "케어포 등록", "케어포 지점", "수급자명", "수급현황", "매칭 연락처", "급여개시일", "퇴소일"])
    hfill = PatternFill("solid", fgColor="4472C4")
    match_fill = PatternFill("solid", fgColor="C6EFCE")
    for c in ws[1]:
        c.fill = hfill
        c.font = Font(bold=True, color="FFFFFF")
    for m, hits in results:
        if hits:
            pt, label = hits[0]
            ws.append([m["center"], m["yearmonth"], m["consult_date"], m["phone"], m["admitted"],
                       "Y", pt["branch"], pt["name"], pt["status"], label, pt["start"], pt["quit"]])
            for c in ws[ws.max_row]:
                c.fill = match_fill
        else:
            ws.append([m["center"], m["yearmonth"], m["consult_date"], m["phone"], m["admitted"],
                       "N", "", "", "", "", "", ""])
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = center
    for i, w in enumerate([12, 12, 12, 14, 12, 10, 11, 11, 10, 14, 12, 12], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    out = Path(__file__).resolve().parent / "상담공지_엑셀" / f"케어포대조_{today:%Y%m%d}.xlsx"
    wb.save(out)
    print(f"\n결과 저장: {out}")

    # 4) 드라이브 업로드(고정 파일, 링크 불변) + 슬랙 링크 공지
    if not args.no_slack:
        from publish_excel import google_token, find_or_create_folder, upload_file, send_slack, ROOT_FOLDER
        token = google_token()
        root_id = find_or_create_folder(token, ROOT_FOLDER)
        info = upload_file(token, out, root_id, "케어포대조_상담미입력.xlsx")
        link = info["webViewLink"]
        print("드라이브 업로드 완료")

        weekday = "월화수목금토일"[today.weekday()]
        msg = {
            "text": f"🔍 케어포 대조 결과 {today.strftime('%Y.%m.%d')}({weekday})",
            "blocks": [
                {"type": "header", "text": {"type": "plain_text", "text": "🔍 상담시트 미입력 × 케어포 대조", "emoji": True}},
                {"type": "context", "elements": [{"type": "mrkdwn",
                    "text": f"{today.strftime('%Y.%m.%d')}({weekday}) · 케어포 1-7 수급자 현황(연간) 연락처 대조"}]},
                {"type": "section", "text": {"type": "mrkdwn",
                    "text": f"미입력 *{len(miss)}건* 중 케어포 수급자 등록 *{len(matched)}건* — "
                            f"입소까지 완료된 상담이므로 상담시트 입력이 꼭 필요합니다.\n"
                            f"📎 <{link}|대조 결과 엑셀 열기/다운로드>"}},
                {"type": "context", "elements": [{"type": "mrkdwn",
                    "text": "🔒 보기 전용 · 초록색 행 = 케어포 등록 확인 건 (수급자명·지점·개시일 포함)"}]},
            ],
        }
        send_slack(msg)
        print("전송 완료 → 슬랙 (webhook)")


if __name__ == "__main__":
    main()
