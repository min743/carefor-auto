# -*- coding: utf-8 -*-
"""
상담 공지 엑셀 생성 — 지점별 파일 + 전체 통합본

내용 (파일당 시트 2개):
  · 신규상담 미입력  : 신규상담 세부사항에서 '상담시트 입력 여부'=N (2026년 5월~)
  · 상담 대기명단    : 센터별 상담 대기 명단 (아웃콜 차수 + 기한 경과일)

실행:
  py -X utf8 consult_excel.py
  → 상담공지_엑셀/YYYY-MM-DD/ 폴더에 센터별 4개 + 전체본 1개 생성
"""
from __future__ import annotations

import sys
from datetime import date
from pathlib import Path

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import consult_report as cr
import waitlist_report as wr

OUT_ROOT = Path(__file__).resolve().parent / "상담공지_엑셀"

FONT_NAME = "맑은 고딕"
BASE_SZ = 13
HEADER_FILL = PatternFill("solid", fgColor="4472C4")
HEADER_FONT = Font(name=FONT_NAME, size=BASE_SZ, bold=True, color="FFFFFF")
BASE_FONT = Font(name=FONT_NAME, size=BASE_SZ)
# 좌측정렬로 볼 긴 텍스트 열(제목 기준) — 잘림·가독성 위해 wrap+top+left
LEFT_COLS = {"AI 요약", "첫 상담일"}
URGENT_FILL = PatternFill("solid", fgColor="FFC7CE")   # 입소완료 미입력 / 기한 지남
TODAY_FILL = PatternFill("solid", fgColor="FFEB9C")    # 오늘 예정

# 맨 뒤 '제외 ✔' 열: 지점이 주간보호 아닌 건에 ✔ 선택(클릭) → 다음 발송 때 제외번호 자동 등록.
MISS_COLS = ["센터명", "구분", "연월", "해당 주차", "상담일자", "급여개시일자", "고객 번호", "입소 여부",
             "케어포 등록", "케어포 지점", "수급자명", "수급현황", "케어포 개시일", "AI 요약", "제외 ✔"]
WAIT_COLS = ["센터명", "아웃콜 차수", "예정일자", "기한 경과(일)", "연락처", "첫 상담일"]
SUMMARY_COLS = ["센터", "신규상담(누적)", "시트 미입력", "미입력률", "⚠️ 입소완료 미입력", "당월 미입력",
                "상담 대기", "기한 지남"]


def _text_width(text: str) -> float:
    """한 줄에 들어오는 최소 열너비(한글=2.1, 그 외=1.15 폭 + 여백)."""
    w = sum(2.1 if ord(ch) > 0x2000 else 1.15 for ch in str(text or ""))
    return w + 2.8  # 좌우 여백


def _style_sheet(ws, widths: list[int], compact: bool = False) -> None:
    """compact=True: 헤더 글씨를 작게(11pt) + 헤더기준 열너비를 좁혀 시트 전체 폭 축소
    (열 많은 '신규상담 미입력' 시트용)."""
    headers = [c.value for c in ws[1]]
    left_idx = {i for i, h in enumerate(headers) if h in LEFT_COLS}
    hdr_font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF") if compact else HEADER_FONT
    hdr_scale = 0.80 if compact else 1.0   # 헤더 11pt는 글자가 작아 더 좁게 담김
    # 열너비 = max(바닥, 헤더+필터화살표, 데이터 최댓값) — 긴 텍스트열(LEFT_COLS)만 상한 적용
    FILTER_PAD = 3.2  # 구글시트/엑셀 헤더의 필터 드롭다운(⇟) 화살표가 먹는 폭
    for i in range(len(headers)):
        col = list(ws.iter_cols(min_col=i + 1, max_col=i + 1))[0]
        data_need = max([_text_width(c.value) for c in col[1:] if c.value is not None] or [0])
        hdr_need = _text_width(headers[i]) * hdr_scale + FILTER_PAD
        base = (5.5 if compact else (widths[i] if i < len(widths) else 10))  # compact은 내용이 폭을 결정
        eff = max(base, data_need, hdr_need)
        if i in left_idx:                 # AI 요약 등 긴 텍스트는 상한 두고 wrap 유지
            eff = min(eff, 34 if compact else max(base, 42))
        ws.column_dimensions[get_column_letter(i + 1)].width = eff
        if i < len(widths):
            widths[i] = eff
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    ws.row_dimensions[1].height = 22 if compact else 24

    center = Alignment(horizontal="center", vertical="center", wrap_text=False)   # 한 줄
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)   # 긴 텍스트만 줄바꿈
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = BASE_FONT
            cell.alignment = left_wrap if (cell.column - 1) in left_idx else center
    _autofit_row_heights(ws, widths, headers)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def _autofit_row_heights(ws, widths: list[int], headers: list) -> None:
    """wrap_text 셀이 눌려 잘려 보이는 문제 해결 — 각 행에 필요한 줄 수를 계산해 높이를 넉넉히 지정.
    (엑셀은 wrap+병합/자동높이 조합에서 자동계산을 안 해줘 '정렬' 눌러야 풀리므로 미리 넣어둔다)"""
    line_px = 18                                  # 12pt 한 줄 대략 높이(px)
    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            v = cell.value
            if v is None:
                continue
            col_w = widths[cell.column - 1] if cell.column - 1 < len(widths) else 12
            cap = max(1, int(col_w / 1.35))       # 열너비로 대략 담기는 글자 수(한글 기준 보수적)
            for seg in str(v).split("\n"):
                seg_len = sum(2 if ord(ch) > 0x2000 else 1 for ch in seg)  # 한글 폭 2 가중
                max_lines = max(max_lines, -(-seg_len // (cap * 2)) or 1)
        ws.row_dimensions[row[0].row].height = min(160, max(18, max_lines * line_px + 4))


def add_summary_sheet(wb: Workbook, srows: list[dict]) -> None:
    ws = wb.create_sheet("요약")
    ws.append(SUMMARY_COLS)
    for s in srows:
        ws.append([s["center"], s["total"], s["miss"], s["rate"],
                   s["urgent"], s["recent"], s["wait"], s["overdue"]])
        if s["center"] == "합계":
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)
    _style_sheet(ws, [13, 14, 12, 10, 17, 11, 10, 10])


def add_miss_sheet(wb: Workbook, rows: list[dict], ym: str, carefor_lookup=None) -> None:
    ws = wb.create_sheet("신규상담 미입력")
    ws.append(MISS_COLS)
    for r in rows:
        if r["admitted"] == "Y":
            kind = "⚠️입소완료"
        elif r["yearmonth"] == ym:
            kind = "당월"
        else:
            kind = "이전"
        # 케어포 수급자 대조 (마지막 다운로드본 기준)
        cf = carefor_lookup(r["phone"]) if carefor_lookup else None
        if cf:
            pt, _label = cf
            cf_cols = ["Y", pt["branch"], pt["name"], pt["status"], pt["start"]]
        else:
            cf_cols = ["N" if carefor_lookup else "", "", "", "", ""]
        ws.append([r["center"], kind, r["yearmonth"], r["week"], r["consult_date"],
                   r["start_date"], r["phone"], r["admitted"], *cf_cols, r["summary"], ""])
        if r["admitted"] == "Y":
            for cell in ws[ws.max_row]:
                cell.fill = URGENT_FILL
    _style_sheet(ws, [11, 15, 11, 11, 11, 12, 13, 8, 9, 10, 10, 9, 11, 40, 8], compact=True)
    # 맨 뒤 '제외 ✔' 열: 클릭→✔ 선택(드롭다운) → 다음 발송 때 자동 제외
    ex_col = get_column_letter(len(MISS_COLS))          # 제외 열
    note_col = get_column_letter(len(MISS_COLS) + 1)    # 옆 안내 열
    last = ws.max_row
    for cell in ws[ex_col][1:]:
        cell.fill = TODAY_FILL
    if last >= 2:
        dv = DataValidation(type="list", formula1='"✔"', allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{ex_col}2:{ex_col}{last}")
    ws[f"{ex_col}1"].comment = Comment(
        "주간보호가 아닌 건(요양원 문의·요양보호사 구인·방문요양 등)이면\n"
        "이 칸을 클릭해 ✔ 선택 → 다음 발송 때 자동으로 제외됩니다.", "안내")
    ws[f"{note_col}1"] = "◀ 예시) 요양원 문의·요양보호사 구인·방문요양 등 주간보호 아닌 건은 클릭해 ✔ (다음날 자동 제외)"
    ws[f"{note_col}1"].font = Font(color="C00000", bold=True)
    ws.column_dimensions[note_col].width = 50


def add_wait_sheet(wb: Workbook, items: list[dict]) -> None:
    ws = wb.create_sheet("상담 대기명단")
    ws.append(WAIT_COLS)
    for it in items:
        over = it["overdue"]
        ws.append([it["center"], it["round"], it["due"],
                   over if over is not None else "", it["phone"], it["first"]])
        if over is not None and over > 0:
            for cell in ws[ws.max_row]:
                cell.fill = URGENT_FILL
        elif over == 0:
            for cell in ws[ws.max_row]:
                cell.fill = TODAY_FILL
    _style_sheet(ws, [12, 14, 14, 13, 15, 13])


def make_book(path: Path, srows: list[dict], miss: list[dict], wait: list[dict], ym: str,
              carefor_lookup=None) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    add_summary_sheet(wb, srows)
    add_miss_sheet(wb, miss, ym, carefor_lookup)
    add_wait_sheet(wb, wait)
    wb.save(path)


def _make_carefor_lookup():
    """케어포 수급자 색인 — 케이포 재조회 없이 기존 다운로드 파일만 사용. 없으면 None(컬럼 비움)."""
    try:
        from carefor_phone_check import DL_DIR, parse_report, digits
        files = sorted(DL_DIR.glob("*_수급자현황_연간.xlsx"))
        if not files:
            return None
        patients = []
        for f in files:
            patients += parse_report(f, f.stem.split("_")[0])
        idx = {}
        for pt in patients:
            for label, d in pt["phones"].items():
                idx.setdefault(d, []).append((pt, label))
        print(f"케어포 대조 사용: 수급자 {len(patients)}명 (다운로드본 {len(files)}개)")
        return lambda phone: (idx.get(digits(phone)) or [None])[0]
    except Exception as e:
        print(f"케어포 대조 생략 (데이터 없음/오류): {e}")
        return None


def _summary_row(name: str, grp: list[dict], wait: list[dict], ym: str) -> dict:
    miss = [r for r in grp if r.get("missing")]
    total = len(grp)
    return {
        "center": name,
        "total": total,
        "miss": len(miss),
        "rate": f"{round(len(miss) / total * 100)}%" if total else "-",
        "urgent": sum(1 for r in miss if r["admitted"] == "Y"),
        "recent": sum(1 for r in miss if r["yearmonth"] == ym),
        "wait": len(wait),
        "overdue": sum(1 for it in wait if it["overdue"] and it["overdue"] > 0),
    }


def generate(today: date | None = None) -> tuple[Path, list[Path], list[dict]]:
    """엑셀 파일들 생성. (출력폴더, [전체본, 센터별...], 센터별 요약) 반환."""
    today = today or date.today()

    # 데이터 로드 (공지 스크립트와 동일 소스)
    a_rows = cr.load_rows_from_webhook()  # r['missing'] 자동 주석(번호대조)
    miss_all = sorted((r for r in a_rows if r.get("missing")),
                      key=lambda r: (r["center"], r["consult_date"]))

    w_raw = wr.load_rows()
    wait_all = []
    for row in w_raw[2:]:
        if len(row) >= 9 and row[0].strip():
            due_d = wr.parse_due(row[8])
            wait_all.append({
                "center": wr.parse_center(row[0]),
                "first": row[1].strip(),
                "round": wr.parse_round(row[7]),
                "due": row[8].strip(),
                "overdue": (today - due_d).days if due_d else None,
                "phone": wr.norm_phone(row[4]),
            })
    wait_all.sort(key=lambda it: (it["center"], -(it["overdue"] if it["overdue"] is not None else -999)))

    out_dir = OUT_ROOT / today.isoformat()
    out_dir.mkdir(parents=True, exist_ok=True)

    ym = f"{today.year}년 {today.month:02d}월"
    paths = []

    # 센터별 요약 계산
    summaries = []
    center_data = {}
    for short, full in cr.CENTER_ORDER:
        grp = [r for r in a_rows if r["center"] == full]
        miss = [r for r in miss_all if r["center"] == full]
        # 대기명단 센터명은 '둔산점'/'서구점' 식 — 공백 제거 후 접두 매칭
        wait = [it for it in wait_all if it["center"].replace(" ", "").startswith(short)]
        center_data[full] = (miss, wait)
        summaries.append(_summary_row(full, grp, wait, ym))
    total_summary = _summary_row("합계", a_rows, wait_all, ym)

    carefor_lookup = _make_carefor_lookup()

    # 전체 통합본 (요약 시트에 센터별 + 합계)
    total_path = out_dir / f"전체_상담공지_{today:%Y%m%d}.xlsx"
    make_book(total_path, summaries + [total_summary], miss_all, wait_all, ym, carefor_lookup)
    paths.append(total_path)
    print(f"생성: {total_path.name}  (미입력 {len(miss_all)} / 대기 {len(wait_all)})")

    # 지점별 파일
    for s, (short, full) in zip(summaries, cr.CENTER_ORDER):
        miss, wait = center_data[full]
        p = out_dir / f"{full}_상담공지_{today:%Y%m%d}.xlsx"
        make_book(p, [s], miss, wait, ym, carefor_lookup)
        paths.append(p)
        print(f"생성: {p.name}  (미입력 {len(miss)} / 대기 {len(wait)})")

    print(f"\n저장 위치: {out_dir}")
    return out_dir, paths, summaries


def main():
    generate()


if __name__ == "__main__":
    main()
