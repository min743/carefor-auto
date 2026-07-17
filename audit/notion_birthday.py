# -*- coding: utf-8 -*-
"""노션 생일쿠폰 발송 내역(위펀 거래명세서) → 월별·지점별 생일자 명단.

- 토큰: NOTION_TOKEN 환경변수(클라우드) 또는 keyring notion_token(로컬 선택).
  없으면 조용히 건너뜀 (클라우드 전용 기능).
- 노션 전월 자료는 매월 8일 이후 업로드 → N월 대조는 (N+1)월 8일 이후에만 판정.
- 파일: CC_(둔산·서구) / CCC_(천안·청주오창) 위펀_전체_거래명세서_YYYY년MM월.xlsx
"""
from __future__ import annotations

import io
import os
import re
from datetime import date, datetime

PAGE_ID = "aaf9857c2dab4e88be4fddc595d8ccd5"  # 생일쿠폰 발송 내역 페이지
NOTION_VERSION = "2022-06-28"
FILE_RE = re.compile(r"(CCC?|ccc?)_.*거래명세서.*?(\d{4})년\s*(\d{1,2})월.*\.xlsx", re.I)

_cache: dict | None = None


def _get_token() -> str | None:
    token = os.environ.get("NOTION_TOKEN")
    if token:
        return token
    try:
        from src import credentials
        return credentials.get("notion_token")
    except Exception:
        return None


def _headers(token: str) -> dict:
    return {"Authorization": f"Bearer {token}", "Notion-Version": NOTION_VERSION}


def _walk_blocks(block_id: str, headers: dict, depth: int = 0, out=None) -> list:
    """블록 트리 순회하며 file 블록 수집."""
    import requests
    if out is None:
        out = []
    if depth > 4:
        return out
    cursor = None
    while True:
        url = f"https://api.notion.com/v1/blocks/{block_id}/children?page_size=100"
        if cursor:
            url += f"&start_cursor={cursor}"
        resp = requests.get(url, headers=headers, timeout=30)
        data = resp.json()
        for blk in data.get("results", []):
            t = blk.get("type")
            if t == "file":
                f = blk["file"]
                name = f.get("name") or "".join(x.get("plain_text", "") for x in blk["file"].get("caption", []))
                url_ = (f.get("file") or f.get("external") or {}).get("url", "")
                out.append({"name": name, "url": url_})
            elif t == "child_database":
                _walk_database(blk["id"], headers, out)
            elif blk.get("has_children"):
                _walk_blocks(blk["id"], headers, depth + 1, out)
        cursor = data.get("next_cursor")
        if not cursor:
            break
    return out


def _walk_database(db_id: str, headers: dict, out: list) -> None:
    """데이터베이스 행의 files 속성 + 행 페이지 내부 파일 수집."""
    import requests
    cursor = None
    while True:
        body = {"page_size": 100}
        if cursor:
            body["start_cursor"] = cursor
        resp = requests.post(f"https://api.notion.com/v1/databases/{db_id}/query",
                             headers={**headers, "Content-Type": "application/json"},
                             json=body, timeout=30)
        data = resp.json()
        for page in data.get("results", []):
            for prop in (page.get("properties") or {}).values():
                if prop.get("type") == "files":
                    for f in prop.get("files", []):
                        out.append({"name": f.get("name", ""),
                                    "url": (f.get("file") or f.get("external") or {}).get("url", "")})
            _walk_blocks(page["id"], headers, 3, out)
        cursor = data.get("next_cursor")
        if not cursor:
            break


def _parse_invoice(content: bytes) -> dict:
    """거래명세서 엑셀 → {지점명: [성명...]}"""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(content), read_only=True)
    result: dict[str, list] = {}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        rows = list(ws.iter_rows(values_only=True))
        hdr_i = next((i for i, r in enumerate(rows)
                      if r and "성명" in [str(c) for c in r if c]), None)
        if hdr_i is None:
            continue
        hdr = [str(c) if c else "" for c in rows[hdr_i]]
        i_branch = next((i for i, h in enumerate(hdr) if "이벤트명" in h), 2)
        i_name = next((i for i, h in enumerate(hdr) if "성명" in h), 3)
        for r in rows[hdr_i + 1:]:
            if not r or r[0] is None:
                continue
            branch = str(r[i_branch] or "").strip()
            name = str(r[i_name] or "").strip()
            if branch and re.match(r"^[가-힣]{2,4}$", name):
                result.setdefault(branch.replace(" ", ""), []).append(name)
    return result


def fetch_birthdays(progress_cb=print) -> dict | None:
    """{'YYYY-MM': {지점명(공백제거): [성명...]}} — 토큰 없으면 None."""
    global _cache
    if _cache is not None:
        return _cache
    token = _get_token()
    if not token:
        return None
    try:
        import requests
        headers = _headers(token)

        # 1) 검색 API로 '생일쿠폰' 관련 페이지/DB를 찾아 우선 수색, 실패 시 대시보드 전체
        roots = []
        try:
            resp = requests.post("https://api.notion.com/v1/search",
                                 headers={**headers, "Content-Type": "application/json"},
                                 json={"query": "생일쿠폰", "page_size": 20}, timeout=30)
            hits = resp.json().get("results", [])
            titles = []
            for h in hits:
                t = ""
                if h["object"] == "page":
                    for prop in (h.get("properties") or {}).values():
                        if prop.get("type") == "title":
                            t = "".join(x.get("plain_text", "") for x in prop["title"])
                elif h["object"] == "database":
                    t = "".join(x.get("plain_text", "") for x in h.get("title", []))
                titles.append(t)
                roots.append((h["object"], h["id"], t))
            progress_cb(f"  노션 검색 '생일쿠폰' → {len(roots)}건: {titles[:5]}")
        except Exception:
            pass

        files = []
        for obj, rid, _t in roots:
            if obj == "database":
                _walk_database(rid, headers, files)
            else:
                _walk_blocks(rid, headers, 0, files)
        if not files:
            files = _walk_blocks(PAGE_ID, headers)
        monthly: dict[str, dict] = {}
        n_parsed = 0
        for f in files:
            m = FILE_RE.search(f.get("name") or "")
            if not m or not f.get("url"):
                continue
            ym = f"{m.group(2)}-{int(m.group(3)):02d}"
            try:
                content = requests.get(f["url"], timeout=60).content
                parsed = _parse_invoice(content)
                n_parsed += 1
            except Exception:
                continue
            dst = monthly.setdefault(ym, {})
            for br, names in parsed.items():
                dst.setdefault(br, []).extend(names)
        progress_cb(f"  노션 생일쿠폰: 파일 {len(files)}개 중 명세서 {n_parsed}개 파싱, {len(monthly)}개월")
        if n_parsed == 0 and files:
            # 진단: 패턴 미일치 시 엑셀류 파일명 샘플 출력 (업무 문서명만 — 개인정보 제외 위해 xlsx 한정)
            samples = [f["name"][:60] for f in files if str(f.get("name", "")).lower().endswith(".xlsx")][:10]
            progress_cb(f"  [진단] xlsx 파일명 샘플: {samples}")
        _cache = monthly
        return monthly
    except Exception as e:
        progress_cb(f"  노션 생일쿠폰 조회 실패(건너뜀): {e}")
        return None


def _floor_ym(opened: str | None, cutoff: str | None) -> str | None:
    """대조 시작 월 'YYYY-MM' = max(기관 지정일, 평가기간 시작). 둘 다 없으면 None(제한 없음).

    ★ 2026-07-17 이전엔 이 필터가 없어 노션에 있는 '모든 달'을 돌았다. 그 결과:
      · 서구점(개소 2025.03) 미지급 의심 165건 중 76건이 2024-02~2025-02 — 지점이
        존재하지도 않던 시기다. 지점은 그 76건을 헛으로 확인해야 했다.
      · 천안(2024.06 개소)·청주(2024.08 개소) 의 "29개월 일치(2024-02~2026-06)" 도
        사실이 아니었다. 개소 전 4~6개월은 '대조해서 맞은' 게 아니라 대상이 아니었는데
        커버리지를 부풀려 말했다.
      branch_pages.py:1132 이 이미 같은 개념(eff = max(cut, opened))을 쓰고 있었는데
      생일쿠폰만 안 탔다.
    """
    ds = []
    for s in (opened, cutoff):
        if not s:
            continue
        try:
            ds.append(datetime.strptime(s.strip(), "%Y.%m.%d").date())
        except (ValueError, AttributeError):
            continue
    if not ds:
        return None
    eff = max(ds)
    return f"{eff.year:04d}-{eff.month:02d}"


def compare(branch_name: str, birthday_log: dict, today: date | None = None,
            progress_cb=print, opened: str | None = None,
            cutoff: str | None = None) -> tuple[list, list] | None:
    """노션 생일자 vs 케어포 대장 지급 대조.

    opened: 기관 지정일자 'YYYY.MM.DD' (9-1 화면). 개소 전 달은 대조 대상이 아니다.
    cutoff: 평가기간 시작 'YYYY.MM.DD'. 평가와 무관한 달까지 지적할 이유가 없다.
    둘 다 안 주면 종전처럼 전 기간을 돈다(호출부가 못 넘기는 경우의 안전한 기본값).

    반환: (미지급 의심 ["YYYY-MM 이름", ...], 판정 월 목록) / 토큰 없으면 None
    """
    today = today or date.today()
    data = fetch_birthdays(progress_cb)
    if data is None:
        return None
    key = branch_name.replace(" ", "")
    floor = _floor_ym(opened, cutoff)
    missing, months, skipped = [], [], 0
    for ym in sorted(data):
        if floor and ym < floor:
            skipped += 1          # 개소 전·평가기간 전 — 대조 대상 아님
            continue
        y, m = int(ym[:4]), int(ym[5:7])
        # N월 자료는 (N+1)월 8일 이후에만 판정
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        if today < date(ny, nm, 8):
            continue
        expected = [n for br, names in data[ym].items() if key in br or br in key for n in names]
        # 해당 월 생일자가 없어도 대조는 수행한 것으로 기록 (지급 없음 = 정상)
        months.append(ym)
        given = set(birthday_log.get(ym, []))
        for n in expected:
            if n not in given:
                missing.append(f"{ym} {n}")
    if skipped:
        progress_cb(f"  노션 생일쿠폰: {floor} 이전 {skipped}개월 제외(개소·평가기간 전)")
    return missing, months
