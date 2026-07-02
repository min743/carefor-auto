# 6개 추가 대조 규칙 분석 → 지점별 엑셀에 '추가대조' 시트 생성
# R1: 낙상 배변>=1 → 화장실 완전자립 불가
# R2: 욕창 영양상태 !=4 → 욕구 영양상태 '양호' 불가
# R3: 욕창 움직임 <=3 → 일어나/옮겨 완전자립 불가
# R4: 낙상 정신상태>=1 → 일어나/옮겨 완전자립 불가
# R5: 낙상 합계>=11 → 옮겨앉기 완전자립 불가
# R6: 욕창 습기여부 !=4 → 화장실 완전자립 불가
import json
import sys
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE = r"C:\Users\alsgm\Desktop\클로드코드"
BRANCHES = {
    "둔산점": (rf"{BASE}\carefor-auto\audit_results\둔산점.json", rf"{BASE}\둔산점_낙상위험도_욕구사정_대조결과.xlsx", "2024.01.01"),
    "서구점": (rf"{BASE}\carefor-auto\audit_results\서구점.json", rf"{BASE}\서구점_낙상위험도_욕구사정_대조결과.xlsx", "2025.01.01"),
    "천안점": (rf"{BASE}\carefor-auto\audit_results\천안점.json", rf"{BASE}\천안점_낙상위험도_욕구사정_대조결과.xlsx", "2024.05.31"),
    "청주 오창점": (rf"{BASE}\carefor-auto\audit_results\청주 오창점.json", rf"{BASE}\낙상위험도_욕구사정_대조결과.xlsx", "2024.07.31"),
}

toN = lambda d: int(str(d).replace(".", ""))


def pair_needs(needs, date):
    eff = None
    for nn in sorted(needs, key=lambda x: toN(x["date"])):
        if toN(nn["date"]) <= toN(date):
            eff = nn
    if eff is None:
        eff = next((nn for nn in sorted(needs, key=lambda x: toN(x["date"])) if toN(nn["date"]) > toN(date)), None)
    return eff


def verdict(cond, needs_val, bad_val, ok_label="일치"):
    """cond=True 이고 needs_val==bad_val 이면 불일치."""
    if not cond:
        return "대상아님"
    if needs_val in ("?", "", None, "실패"):
        return "확인필요(미체크)"
    return "불일치" if needs_val == bad_val else ok_label


def analyze_branch(branch, src, cutoff):
    d = json.load(open(src, encoding="utf-8"))
    rows = []
    for p in d["raw"]:
        if p.get("err"):
            continue
        name, status = p["name"], p.get("status", "")
        needs = p.get("needs", [])

        # 낙상 기반 규칙 (R1, R4, R5)
        for f in p.get("falls", []):
            if toN(f["date"]) < toN(cutoff):
                continue
            eff = pair_needs(needs, f["date"])
            nd = eff["date"] if eff else ""
            toilet = (eff or {}).get("toilet", "?")
            sit = (eff or {}).get("sit", "?")
            tr = (eff or {}).get("tr", "?")
            bv, ms, total = f.get("bv", -9), f.get("ms", -9), f.get("total", -9)
            if bv is not None and bv >= 1:
                rows.append([name, status, "R1 배변→화장실", f["date"], f"배변 {bv}점", nd, f"화장실: {toilet}",
                             verdict(True, toilet, "완전자립")])
            if ms is not None and ms >= 1:
                v_sit = verdict(True, sit, "완전자립")
                v_tr = verdict(True, tr, "완전자립")
                v = "불일치" if "불일치" in (v_sit, v_tr) else ("확인필요(미체크)" if "확인필요(미체크)" in (v_sit, v_tr) else "일치")
                rows.append([name, status, "R4 정신상태→일어나/옮겨", f["date"], f"정신상태 {ms}점", nd,
                             f"일어나: {sit} / 옮겨: {tr}", v])
            if total is not None and total >= 11:
                rows.append([name, status, "R5 합계11점↑→옮겨앉기", f["date"], f"합계 {total}점", nd, f"옮겨: {tr}",
                             verdict(True, tr, "완전자립")])

        # 욕창 기반 규칙 (R2, R3, R6)
        for s in p.get("sores", []):
            if toN(s["date"]) < toN(cutoff):
                continue
            sc = s.get("scores") or {}
            def score_of(*labels):
                for lb in labels:
                    for k, v in sc.items():
                        if k.startswith(lb):
                            return v.get("score")
                return None
            nut_s = score_of("영양상태", "영양")
            mob_s = score_of("움직임")
            wet_s = score_of("습기")
            eff = pair_needs(needs, s["date"])
            nd = eff["date"] if eff else ""
            nut_n = (eff or {}).get("nutrition", "?")
            sit = (eff or {}).get("sit", "?")
            tr = (eff or {}).get("tr", "?")
            toilet = (eff or {}).get("toilet", "?")
            if nut_s is not None and nut_s != 4:
                rows.append([name, status, "R2 영양상태→욕구영양", s["date"], f"영양상태 {nut_s}점", nd,
                             f"욕구 영양: {nut_n}", verdict(True, nut_n, "양호")])
            if mob_s is not None and mob_s <= 3:
                v_sit = verdict(True, sit, "완전자립")
                v_tr = verdict(True, tr, "완전자립")
                v = "불일치" if "불일치" in (v_sit, v_tr) else ("확인필요(미체크)" if "확인필요(미체크)" in (v_sit, v_tr) else "일치")
                rows.append([name, status, "R3 움직임→일어나/옮겨", s["date"], f"움직임 {mob_s}점", nd,
                             f"일어나: {sit} / 옮겨: {tr}", v])
            if wet_s is not None and wet_s != 4:
                rows.append([name, status, "R6 습기여부→화장실", s["date"], f"습기 {wet_s}점", nd,
                             f"화장실: {toilet}", verdict(True, toilet, "완전자립")])
    return rows


def write_sheet(xlsx, branch, cutoff, rows):
    wb = load_workbook(xlsx)
    SHEET = "추가대조"
    if SHEET in wb.sheetnames:
        del wb[SHEET]
    ws = wb.create_sheet(SHEET)

    head_font = Font(name="맑은 고딕", bold=True, size=11)
    base_font = Font(name="맑은 고딕", size=10)
    red_font = Font(name="맑은 고딕", size=10, color="C00000", bold=True)
    red_fill = PatternFill("solid", start_color="FCE4EC")
    head_fill = PatternFill("solid", start_color="D9E1F2")
    thin = Border(*[Side(style="thin")] * 4)

    n_disc = sum(1 for r in rows if r[7] == "불일치")
    per_rule = {}
    for r in rows:
        if r[7] == "불일치":
            per_rule[r[2].split(" ")[0]] = per_rule.get(r[2].split(" ")[0], 0) + 1

    ws["A1"] = f"추가 대조 6개 규칙 — {branch} (기준일 {cutoff}~)"
    ws["A1"].font = Font(name="맑은 고딕", bold=True, size=13)
    ws["A2"] = ("R1 배변≥1→화장실 | R2 욕창영양≠4→욕구영양'양호'불가 | R3 움직임≤3→일어나/옮겨 | "
                "R4 정신상태≥1→일어나/옮겨 | R5 낙상합계≥11→옮겨 | R6 습기≠4→화장실")
    ws["A2"].font = base_font
    ws["A3"] = f"전체 {len(rows)}건 중 불일치 {n_disc}건 — " + ", ".join(f"{k} {v}건" for k, v in sorted(per_rule.items())) \
               + f" | 생성 {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws["A3"].font = head_font

    hdrs = ["수급자명", "현황", "규칙", "평가일", "평가 점수", "욕구사정일", "욕구사정 값", "판정"]
    for c, h in enumerate(hdrs, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = head_font
        cell.fill = head_fill
        cell.border = thin
        cell.alignment = Alignment(horizontal="center")
    r = 5
    for row in rows:
        r += 1
        is_disc = row[7] == "불일치"
        for c, v in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = red_font if is_disc else base_font
            if is_disc:
                cell.fill = red_fill
    for col, w in zip("ABCDEFGH", [13, 8, 24, 12, 13, 12, 26, 15]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = "A6"
    ws.auto_filter.ref = f"A5:H{r}"
    wb.save(xlsx)
    return n_disc, per_rule, len(rows)


if __name__ == "__main__":
    for branch, (src, xlsx, cutoff) in BRANCHES.items():
        try:
            rows = analyze_branch(branch, src, cutoff)
            n_disc, per_rule, total = write_sheet(xlsx, branch, cutoff, rows)
            disc_people = len({r[0] for r in rows if r[7] == "불일치"})
            print(f"[{branch}] {total}건 검사 → 불일치 {n_disc}건/{disc_people}명 | " +
                  ", ".join(f"{k}:{v}" for k, v in sorted(per_rule.items())))
        except Exception as e:
            print(f"[{branch}] 실패: {e}")
